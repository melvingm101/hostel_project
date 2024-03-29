# Create your views here.
import openpyxl
from openpyxl.styles import Border, Alignment, Font
from copy import copy
import sqlite3

from django import forms
from django.shortcuts import render
from students.forms import LoginForm
from django.contrib.auth.decorators import login_required
from students.models import Student, Charges, Other_Charges
import datetime


@login_required()
def front_page(request):
    normal = []
    if request.method == "POST" and request.FILES['excel_file']:
        # connection = sqlite3.connect('students.db')
        # cursor = connection.cursor()
        #
        # print("Table created!")
        # sql_create = """ CREATE TABLE IF NOT EXISTS students(
        #     NUMBER integer PRIMARY KEY AUTOINCREMENT,
        #     ROOMNO VARCHAR(4),
        #     BLOCK VARCHAR(3),
        #     STUDENT VARCHAR(30),
        #     MOBILE VARCHAR(20),
        #     ID VARCHAR(20)
        # ) """
        #
        # cursor.execute(sql_create)
        # print("Table created!")


        #sql_insert = """ INSERT INTO students VALUES("312", "A", "Melvin George Mathew", "0509940039", "2015A7PS0197U")"""
        #cursor.execute(sql_insert)
        #print("You just got posted son!")

        excel_files = request.FILES["excel_file"]
        wb = openpyxl.load_workbook(excel_files)
        worksheet = wb["Sheet1"]
        # #print(worksheet.get_highest_column())
        # for row in worksheet.rows:
        #     normal.append(row)
        # for rows in normal:
        #     print(rows[1].value)
        #     sql_insert = "INSERT INTO students(ROOMNO, BLOCK, STUDENT, MOBILE, ID) VALUES('{0}', '{1}', '{2}', '{3}', '{4}')".format(rows[0].value, rows[1].value, rows[2].value, rows[3].value, rows[4].value)
        #     cursor.execute(sql_insert)
        #     print("Data added!")
        #
        # cursor.execute("SELECT * FROM students")
        # res = cursor.fetchall()
        # connection.commit()
        # connection.close()
        # return render(request, 'homepage.html', {'res': res})

        #connection = sqlite3.connect('students.db')
        #cursor = connection.cursor()
        # sql_create = """ CREATE TABLE IF NOT EXISTS students(
        #             NUMBER integer PRIMARY KEY AUTOINCREMENT,
        #             ROOMNO VARCHAR(4),
        #             BLOCK VARCHAR(3),
        #             STUDENT VARCHAR(30),
        #             MOBILE VARCHAR(20),
        #             ID VARCHAR(20)
        #         ) """
        #
        # cursor.execute(sql_create)
        stud_rows=list(worksheet.rows)
        for row in stud_rows[1:]:
            args=[cell.value for cell in row]
            student=Student(name=args[0],id_no=args[1],room_no=args[2],joining_date=args[3],leaving_date=args[4],block=args[5],mob_no=args[6],email=args[7],date_added=datetime.datetime.now())
            charges=Charges(laundry_charges=args[8],lost_key_charges=args[9],breakage_charges=args[10],guest_charges=args[11])
            other_charges=Other_Charges()
            other_charges.save()
            charges.other_charges=other_charges
            charges.save()
            student.charges=charges
            student.save()
        # cursor.execute("SELECT * FROM students")
        # res = cursor.fetchall()
        # connection.commit()
        # connection.close()
        # return render(request, 'homepage.html', {'res': res})
    res=Student.objects.all()
    return render(request, 'homepage.html',{'res':res})


def login_page(request):
    return render(request, 'login.html')

# def guest(request):
#     if request.method == "POST" and request.FILES['excel_file']:
#         excel_files = request.FILES["excel_file"]
#         wb = openpyxl.load_workbook(excel_files)
#         worksheet = wb["Sheet1"]
#         guest_rows = list(worksheet.rows)
#         for row in guest_rows[1:]:
#             args = [cell.value for cell in row]
#             guest = Guest(name=args[0],category=args[1], joining_date=args[2], leaving_date=args[3], room_no=args[4], mob_no=args[5], email=args[6], charges=args[7],date_added=datetime.datetime.now())
#             guest.save()
#
#     res=Guest.objects.all()
#     return render(request, '../guests/templates/guests.html', {'res':res})


@login_required()
def insertStudents(request):
    print("This is the insert Excel page")
    connection = sqlite3.connect('students.db')
    cursor = connection.cursor()
    if request.method == "POST":
        data = request.POST.copy()
        student=Student()
        student.room_no = data.get('room_no')
        student.block = data.get('block')
        student.name = data.get('name')
        student.mob_no = data.get('mob_no')
        student.id_no = data.get('id_no')
        print(student.room_no)
        student.save()
        #sql_insert_single = "INSERT INTO students(ROOMNO, BLOCK, STUDENT, MOBILE, ID) VALUES('{0}', '{1}', '{2}', '{3}', '{4}')".format(room_no, block, student, mobile, id)
        #cursor.execute(sql_insert_single)
        connection.commit()
        connection.close()
    return render(request, 'insertStudentExcel.html')


# @login_required()
# def insert_guest(request):
#     return render(request, '../guests/templates/insertGuestExcel.html')

def login(request):
    if request.method == "POST":
        form = LoginForm(request.POST)
        print("YES")
        if form.is_valid():
            user = form.cleaned_data['user']

    else:
        form = LoginForm()

    return render(request, 'login.html', {"form": form})

def upload_excel(request):
    print("Boss mode")

def studentDetail(request, student_id):
    student=Student.objects.get(pk=student_id)
    return render(request, 'studentDetail.html', {'student':student})

# def guest_detail(request,guest_id):
#     guest=Guest.objects.get(pk=guest_id)
#     return render(request, '../guests/templates/GuestDetail.html', {'guest':guest})

def generateStudentSummary(request):
    wb=openpyxl.Workbook()
    ws=wb.active
    ws['D1']='BITS Pilani, Dubai Campus'
    ws.merge_cells('D1:F1')
    heading=ws['D1']
    heading.alignment=Alignment(horizontal='center', vertical='center')
    heading.font=Font(bold=True)
    ws['D3']='Students Report'
    ws.merge_cells('D3:F3')
    subheading=ws['D3']
    subheading=copy(heading)

    ws['A5']='S.No'
    ws['B5']='Database ID'
    ws['C5']='Name'
    ws['D5']='ID Number'
    ws['E5']='Room Number'
    ws['F5']='Joining Date'
    ws['G5']='Leaving Date'
    ws['H5']='Block'
    ws['I5']='Mobile Number'
    ws['J5']='Email'
    ws['K5']='Total Amount Due'

    row=ws.row_dimensions[5]
    row.font=Font(bold=True)

    students=list(Student.objects.all())
    i=1
    for stud in students:
        row=[i,stud.id,stud.name,stud.id_no,stud.room_no,stud.joining_date,stud.leaving_date,stud.block,stud.mob_no,stud.email]
        charges=Charges.objects.get(pk=stud.charges.id)
        amount=charges.lost_key_charges+charges.breakage_charges+charges.guest_charges+charges.laundry_charges
        row.append(amount)
        ws.append(row)
        i=i+1

    wb.save('testStudentsReport.xlsx')

    return render(request,'homepage.html')


class UploadFileForm(forms.Form):
    file = forms.FileField()

class NameForm(forms.Form):
    room_no = forms.Field
    block = forms.CharField(max_length=100)
    student = forms.CharField(max_length=100)
    mobile = forms.CharField(max_length=100)
    id = forms.CharField(max_length=100)


