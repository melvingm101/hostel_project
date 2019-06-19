from django.shortcuts import render
import openpyxl
from openpyxl.styles import Border, Alignment, Font
from copy import copy
import sqlite3

from django import forms
from django.shortcuts import render
from students.forms import LoginForm
from django.contrib.auth.decorators import login_required
from guests.models import Guest
import datetime

# Create your views here.

def guestHome(request):
    if request.method == "POST" and request.FILES['excel_file']:
        excel_files = request.FILES["excel_file"]
        wb = openpyxl.load_workbook(excel_files)
        worksheet = wb["Sheet1"]
        guest_rows = list(worksheet.rows)
        for row in guest_rows[1:]:
            args = [cell.value for cell in row]
            guest = Guest(name=args[0],category=args[1], joining_date=args[2], leaving_date=args[3], room_no=args[4], mob_no=args[5], email=args[6], charges=args[7],date_added=datetime.datetime.now())
            guest.save()

    res=Guest.objects.all()
    return render(request, 'guests.html', {'res':res})


@login_required()
def insert_guest(request):
    return render(request, 'insertGuestExcel.html')

def guest_detail(request,guest_id):
    guest=Guest.objects.get(pk=guest_id)
    return render(request, 'GuestDetail.html', {'guest':guest})

def generateGuestSummary(request):
    wb=openpyxl.Workbook()
    ws=wb.active
    ws['D1']='BITS Pilani, Dubai Campus'
    ws.merge_cells('D1:F1')
    heading=ws['D1']
    heading.alignment=Alignment(horizontal='center', vertical='center')
    heading.font=Font(bold=True)
    ws['D3']='Guests Report'
    ws.merge_cells('D3:F3')
    subheading=ws['D3']
    subheading=copy(heading)

    ws['A5']='S.No'
    ws['B5']='Database ID'
    ws['C5']='Name'
    ws['D5']='Category'
    ws['E5']='Room Number'
    ws['F5']='Joining Date'
    ws['G5']='Leaving Date'
    ws['H5']='Mobile Number'
    ws['I5']='Email'
    ws['J5']='Total Amount Due'

    row=ws.row_dimensions[5]
    row.font=Font(bold=True)

    guests=list(Guest.objects.all())
    i=1
    for guest in guests:
        row=[i,guest.id,guest.name,guest.category,guest.room_no,guest.joining_date,guest.leaving_date,guest.mob_no,guest.email,guest.charges]
        ws.append(row)
        i=i+1

    wb.save('testGuestsReport.xlsx')

    return render(request,'guests.html')